# -*- coding: utf-8 -*-
"""
Convert the Excel Google review file into Qdrant — the reusable "ingest" function.

This module is the single source of truth for ingestion. It is used by:
  - embedding.py  -> the CLI entry point (python embedding.py)
  - WebApps/api_routes.py -> the POST /api/convert endpoint (button on the /convert page)

The function mirrors what embedding.py used to do as a top-level script:
  1. Read the Excel workbook with openpyxl.
  2. Wrap each review as a LangChain Document (text to embed + structured metadata).
  3. QdrantVectorStore.from_documents() embeds every review with Ollama
     (qwen3-embedding:0.6b), auto-creates/opens the "restaurant_reviews" collection,
     and upserts everything into Qdrant.

Idempotent: deterministic UUID5 ids mean re-running updates the same points
instead of creating duplicates — safe to call from the web page repeatedly.
"""
import threading
import uuid
from pathlib import Path

from openpyxl import load_workbook  # reads .xlsx (Excel) files
from langchain_core.documents import Document  # "text to embed" + "metadata to keep"
from langchain_ollama import OllamaEmbeddings  # turns text into vectors (runs locally via Ollama)
from langchain_qdrant import QdrantVectorStore  # talks to Qdrant (our vector database)

# --- Config ------------------------------------------------------------------
EXCEL_FILE = "ABC123_Restaurant_Kuching_Google_Reviews_100_Rows.xlsx"
# All spreadsheet files live in the spreadsheet_data/ folder next to this file,
# so the path works no matter which directory the script is started from.
SPREADSHEET_DIR = Path(__file__).resolve().parent / "spreadsheet_data"
DEFAULT_EXCEL_PATH = SPREADSHEET_DIR / EXCEL_FILE  # default file to convert
QDRANT_URL = "http://localhost:6333"
COLLECTION_NAME = "restaurant_reviews"
EMBED_MODEL = "qwen3-embedding:0.6b"

# Prevents two /api/convert calls from embedding at the same time. A second
# call simply waits for the first to finish, then does its own upsert.
_LOCK = threading.Lock()


def convert_excel_to_qdrant(
    excel_file: str = str(DEFAULT_EXCEL_PATH),
    qdrant_url: str = QDRANT_URL,
    collection_name: str = COLLECTION_NAME,
    embed_model: str = EMBED_MODEL,
) -> dict:
    """Read the Excel review file, embed each row, and upsert it into Qdrant.

    Returns a summary dict: {"converted": <rows embedded>, "points": <points in Qdrant>}.
    """
    with _LOCK:  # only one conversion runs at a time (see comment above)
        # --- Step 1: read the reviews ----------------------------------------
        # load_workbook opens the .xlsx file; wb.active is the first (usually only) worksheet.
        wb = load_workbook(excel_file)
        # iter_rows(values_only=True) returns each row as a plain tuple of values.
        rows = list(wb.active.iter_rows(values_only=True))
        headers = rows[0]  # first row = column names, e.g. "Review ID", "Review Text"
        documents = []
        for values in rows[1:]:  # every row AFTER the header is one review
            if not values or values[0] is None:  # skip completely empty rows
                continue
            # zip(headers, values) pairs each column name with its value, and dict()
            # turns that into a normal dictionary: row["Review Text"] -> the review text.
            row = dict(zip(headers, values))

            # Qdrant stores payloads as JSON, and JSON cannot hold datetime objects,
            # so convert dates to plain "YYYY-MM-DD" strings.
            for key in ("Review Date", "Visit Date"):
                row[key] = row[key].strftime("%Y-%m-%d") if hasattr(row[key], "strftime") else str(row[key])
            row["Star Rating"] = int(row["Star Rating"])  # "4" -> 4 (so filtering works numerically)

            # A LangChain Document has two parts:
            #   page_content -> the TEXT that gets embedded (the review itself)
            #   metadata     -> structured fields kept alongside, used later for
            #                   filtering/sorting, but NOT embedded.
            documents.append(
                Document(
                    page_content=row["Review Text"],
                    metadata={
                        "review_id": row["Review ID"],
                        "review_date": row["Review Date"],
                        "reviewer_name": row["Reviewer Name"],
                        "star_rating": row["Star Rating"],
                        "visit_date": row["Visit Date"],
                        "reviewer_location": row["Reviewer Location"],
                        "order_type": row["Order Type"],
                        "dish_mentioned": row["Dish Mentioned"],
                    },
                )
            )

        # --- Step 2: embed everything and upsert into Qdrant -----------------
        # uuid5(NAMESPACE_URL, review_id) always produces the SAME id for the same
        # review_id. That keeps the conversion idempotent: re-running it overwrites
        # the same points instead of adding duplicates.
        ids = [str(uuid.uuid5(uuid.NAMESPACE_URL, d.metadata["review_id"])) for d in documents]

        # from_documents() does 3 things for us:
        #   1. embeds every page_content with the Ollama embedding model,
        #   2. auto-creates the collection (vector size is inferred from the model, 1024),
        #   3. upserts all documents (with their metadata) into Qdrant.
        store = QdrantVectorStore.from_documents(
            documents,
            embedding=OllamaEmbeddings(model=embed_model),
            url=qdrant_url,
            collection_name=collection_name,
            ids=ids,
        )

        # Sanity check: how many points does Qdrant actually hold now?
        count = store.client.count(collection_name).count
        return {"converted": len(documents), "points": count}


if __name__ == "__main__":
    # CLI entry point kept here as a convenience:  python converter.py
    summary = convert_excel_to_qdrant()
    print(f"Embedded {summary['converted']} reviews into '{COLLECTION_NAME}' ({summary['points']} points in Qdrant).")
