# -*- coding: utf-8 -*-
"""
Embed the Excel Google review rows into Qdrant — the "ingest" step of the RAG pipeline.

What this script does (3 steps):
  1. Read the Excel workbook with openpyxl.
  2. Wrap each review as a LangChain Document (text to embed + structured metadata).
  3. QdrantVectorStore.from_documents() embeds every review with Ollama
     (qwen3-embedding:0.6b), auto-creates the "restaurant_reviews" collection,
     and uploads everything into Qdrant.

Key concepts (beginner note):
  - Embedding: turning a piece of TEXT into a list of numbers (a "vector") that
    captures its meaning. Similar texts get similar vectors.
  - Vector store: a database that stores those vectors and can later find the
    most similar ones to a new text (chat.py uses it for retrieval).
  - Idempotent: safe to run more than once — running it again with the same data
    updates the existing points instead of creating duplicates.

Run with:  python embedding.py
"""
import datetime
import uuid

from openpyxl import load_workbook  # reads .xlsx (Excel) files
from langchain_core.documents import Document  # LangChain wrapper: "text to embed" + "metadata to keep"
from langchain_ollama import OllamaEmbeddings  # turns text into vectors (runs locally via Ollama)
from langchain_qdrant import QdrantVectorStore  # talks to Qdrant (our vector database)

# --- Config ------------------------------------------------------------------
EXCEL_FILE = "ABC123_Restaurant_Kuching_Google_Reviews_100_Rows.xlsx"
QDRANT_URL = "http://localhost:6333"
COLLECTION_NAME = "restaurant_reviews"
EMBED_MODEL = "qwen3-embedding:0.6b"

# --- Step 1: read the reviews ------------------------------------------------
# load_workbook opens the .xlsx file; wb.active is the first (usually only) worksheet.
wb = load_workbook(EXCEL_FILE)
# iter_rows(values_only=True) returns each row as a plain tuple of values
# (no cell objects), which is easy to work with.
rows = list(wb.active.iter_rows(values_only=True))
headers = rows[0]  # first row = column names, e.g. "Review ID", "Review Text"
documents = []
for values in rows[1:]:  # every row AFTER the header is one review
    if not values or values[0] is None:  # skip completely empty rows
        continue
    # zip(headers, values) pairs each column name with its value, and dict()
    # turns that into a normal dictionary: row["Review Text"] -> the review text.
    row = dict(zip(headers, values))

    # Qdrant stores payloads as JSON, and JSON cannot hold datetime objects,
    # so convert dates to plain "YYYY-MM-DD" strings.
    for key in ("Review Date", "Visit Date"):
        row[key] = row[key].strftime("%Y-%m-%d") if isinstance(row[key], datetime.datetime) else str(row[key])
    row["Star Rating"] = int(row["Star Rating"])  # "4" -> 4 (so filtering works numerically)

    # A LangChain Document has two parts:
    #   page_content -> the TEXT that gets embedded (the review itself)
    #   metadata     -> structured fields kept alongside, used later for
    #                   filtering/sorting, but NOT embedded.
    documents.append(
        Document(
            page_content=row["Review Text"],
            metadata={
                "review_id": row["Review ID"],
                "review_date": row["Review Date"],
                "reviewer_name": row["Reviewer Name"],
                "star_rating": row["Star Rating"],
                "visit_date": row["Visit Date"],
                "reviewer_location": row["Reviewer Location"],
                "order_type": row["Order Type"],
                "dish_mentioned": row["Dish Mentioned"],
            },
        )
    )

# --- Step 2: embed everything and upsert into Qdrant -------------------------
# uuid5(NAMESPACE_URL, review_id) always produces the SAME id for the same
# review_id. That makes the script idempotent: re-running it overwrites the
# same points instead of adding duplicates.
ids = [str(uuid.uuid5(uuid.NAMESPACE_URL, d.metadata["review_id"])) for d in documents]

# from_documents() does 3 things for us:
#   1. embeds every page_content with the Ollama embedding model,
#   2. auto-creates the collection (vector size is inferred from the model, 1024),
#   3. upserts all documents (with their metadata) into Qdrant.
store = QdrantVectorStore.from_documents(
    documents,
    embedding=OllamaEmbeddings(model=EMBED_MODEL),
    url=QDRANT_URL,
    collection_name=COLLECTION_NAME,
    ids=ids,
)

# Sanity check: how many points does Qdrant actually hold now?
count = store.client.count(COLLECTION_NAME).count
print(f"Embedded {len(documents)} reviews into '{COLLECTION_NAME}' ({count} points in Qdrant).")
